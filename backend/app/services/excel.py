import io
import re
from datetime import date, datetime

import openpyxl
from openpyxl.utils import get_column_letter

from app.models.asset import ASSET_STATUS_LABELS

# Single canonical column mapping used for both export and import, per
# docs/feature-spec.md §9 ("one canonical schema, adapters" rather than
# N special-case scripts). Column order here is also the .xlsx export
# column order — import no longer depends on position at all (see
# parse_asset_xlsx), only export does, so this stays append-only for
# anyone reading an exported file by eye.
COLUMNS = [
    ("asset_code", "Mã tài sản"),
    ("name", "Tên tài sản"),
    ("category", "Nhóm thiết bị"),
    ("spec", "Cấu hình kỹ thuật"),
    ("serial_number", "Số Serial"),
    ("manufacturer", "Hãng sản xuất"),
    ("manufacture_year", "Năm sản xuất"),
    ("original_cost", "Nguyên giá (VNĐ)"),
    ("warranty_months", "Bảo hành (tháng)"),
    ("department", "Phòng ban"),
    ("holder", "Người sử dụng"),
    ("location", "Vị trí"),
    ("status", "Tình trạng"),
    ("purchase_source", "Nơi mua"),
    ("notes", "Ghi chú"),
    # Not an Asset column: resolved to/from holder_user_id at the API layer
    # (see assets.py import/export), not passed straight through like the
    # other fields here.
    ("holder_email", "Email người sử dụng"),
    # These three are real, pre-existing Asset columns (see app/models/asset.py)
    # that simply had no Excel column of their own until now — appended
    # rather than inserted, same backward-compatibility reason as holder_email.
    ("year_put_in_use", "Năm đưa vào sử dụng"),
    ("budget_plan_year", "Kế hoạch NS năm"),
    ("budget_actual_year", "Thực hiện NS năm"),
]

_STATUS_LABEL_TO_CODE = {label.lower(): code for code, label in ASSET_STATUS_LABELS.items()}

# Extra known real-world header variants per field, beyond the canonical
# export label itself (always included automatically below) — collected
# from actual historical spreadsheets this app needs to import (TB-VP /
# ĐPHA / TB-Bỏ style sheets), not just from this app's own export format,
# since re-importing a file this app never produced is the whole point.
_FIELD_ALIASES = {
    "name": ["tên vật tư- thiết bị", "tên vật tư thiết bị", "tên thiết bị", "tên tài sản/vật tư"],
    "spec": ["cấu hình kỹ thuật cơ bản"],
    "serial_number": ["serial"],
    "original_cost": ["nguyên giá", "nguyễn giá"],  # "nguyễn" is a common real-world typo for "nguyên"
    "warranty_months": ["bảo hành"],
    "department": ["phòng ban sử dụng"],
    "holder": ["người sử dụng"],
    "year_put_in_use": ["năm sử dụng"],
    "budget_plan_year": ["kế hoạch ns năm"],
    "budget_actual_year": ["thực hiện ns năm"],
}

# Columns that carry no asset data of their own (row numbering, section
# markers) — recognized and skipped rather than captured into extra_fields.
_IGNORED_HEADERS = {"stt", "số thứ tự", "#"}

_SECTION_MARKER_RE = re.compile(r"^[ivxlcdm]+$")


def _normalize(value) -> str:
    """Case/whitespace/newline-insensitive header comparison key. Does NOT
    strip Vietnamese diacritics — tone marks change meaning (and this app's
    real-world header typos, like "nguyễn" for "nguyên", are handled as
    explicit aliases instead, since blanket accent-folding risks merging
    genuinely different words)."""
    if value is None:
        return ""
    text = str(value).replace("\n", " ").replace("\r", " ")
    return re.sub(r"\s+", " ", text).strip().lower()


def _aliases_for(field: str, label: str) -> list[str]:
    return [_normalize(label)] + [_normalize(a) for a in _FIELD_ALIASES.get(field, [])]


_FIELD_ALIAS_LISTS = [(field, _aliases_for(field, label)) for field, label in COLUMNS]


def _match_field(header_norm: str) -> str | None:
    """Best-effort match of a normalized header cell to a known Asset
    field: exact alias match first, then substring containment either
    direction (handles stray suffix words like "CƠ BẢN" or a unit note in
    parentheses that the alias list doesn't spell out verbatim)."""
    if not header_norm:
        return None
    for field, aliases in _FIELD_ALIAS_LISTS:
        if header_norm in aliases:
            return field
    for field, aliases in _FIELD_ALIAS_LISTS:
        for alias in aliases:
            if alias and (alias in header_norm or header_norm in alias):
                return field
    return None


def _detect_header_row(ws, max_scan_rows: int = 20) -> tuple[int, tuple]:
    """Real-world sheets often have title rows above the actual header, and
    sometimes a two-tier grouped header (a section title spanning several
    columns, with the real per-column labels one row below) — scan for the
    row that looks most like a header instead of assuming row 1."""
    best_idx, best_row, best_score = None, None, -1
    for idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=min(max_scan_rows, ws.max_row), values_only=True), start=1
    ):
        score = sum(1 for cell in row if _match_field(_normalize(cell)))
        if score > best_score:
            best_idx, best_row, best_score = idx, row, score
    if best_score < 2:
        raise ValueError(
            "Không tìm thấy dòng tiêu đề phù hợp — vui lòng kiểm tra file có các cột "
            "như Tên tài sản, Mã tài sản, Phòng ban, v.v."
        )
    return best_idx, best_row


def _coerce_int(value):
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(value)
    match = re.search(r"\d+", str(value))
    return int(match.group()) if match else None


def _coerce_float(value):
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    digits = re.sub(r"[^\d.,]", "", str(value)).replace(".", "").replace(",", "")
    return float(digits) if digits.isdigit() else None


def _coerce_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def build_asset_xlsx(assets, holder_emails: dict | None = None) -> bytes:
    """`holder_emails` maps Asset.holder_user_id -> email, resolved by the
    caller (assets.py) since `holder_email` isn't a real Asset column."""
    holder_emails = holder_emails or {}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách tài sản"

    for col_idx, (_, label) in enumerate(COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=label)

    for row_idx, asset in enumerate(assets, start=2):
        for col_idx, (field, _) in enumerate(COLUMNS, start=1):
            if field == "holder_email":
                value = holder_emails.get(asset.holder_user_id)
            else:
                value = getattr(asset, field)
                if field == "status":
                    value = ASSET_STATUS_LABELS.get(value, value)
                elif field == "original_cost" and value is not None:
                    value = float(value)
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_idx in range(1, len(COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def parse_asset_xlsx(content: bytes) -> list[dict]:
    """Parses an uploaded workbook by matching column HEADERS (not
    position) against known Asset fields, so any file whose columns are
    reordered, renamed, missing, or exported by a different system than
    this app entirely can still be imported:
      - a header that matches a known field (by exact or fuzzy alias) fills
        that field for every row;
      - a known field with no matching column in the file is simply left
        empty on every imported row;
      - a header that matches nothing known is preserved per-row under
        `extra_fields`, keyed by its raw text, rather than discarded.
    Raises ValueError only if no header-like row can be found at all.
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    header_row_idx, header_row = _detect_header_row(ws)

    field_by_col: dict[int, str] = {}
    extra_col_labels: dict[int, str] = {}
    matched_fields: set[str] = set()
    for col_idx, cell in enumerate(header_row):
        norm = _normalize(cell)
        if not norm or norm in _IGNORED_HEADERS:
            continue
        field = _match_field(norm)
        if field and field not in matched_fields:
            field_by_col[col_idx] = field
            matched_fields.add(field)
        elif not field:
            extra_col_labels[col_idx] = str(cell).strip()

    rows = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not any(row):
            continue

        # Section/category divider rows (e.g. "I", "II" with just a group
        # label like "Máy tính xách tay Laptop" and nothing else) — real
        # row numbers are 1, 2, 3…, so a bare roman numeral in the first
        # column means this is a heading, not an asset.
        first_cell = _normalize(row[0]) if row else ""
        if first_cell and _SECTION_MARKER_RE.match(first_cell):
            continue

        mapped_raw = {col_idx: (row[col_idx] if col_idx < len(row) else None) for col_idx in field_by_col}
        if not any(v not in (None, "") for v in mapped_raw.values()):
            continue  # no recognizable asset data on this row at all

        record: dict = {}
        for col_idx, field in field_by_col.items():
            value = mapped_raw[col_idx]
            if isinstance(value, str):
                value = value.strip() or None
            if field == "status":
                value = _STATUS_LABEL_TO_CODE.get(str(value or "").strip().lower(), "dang_su_dung")
            elif field in ("manufacture_year", "warranty_months", "budget_plan_year", "budget_actual_year"):
                value = _coerce_int(value)
            elif field == "original_cost":
                value = _coerce_float(value)
            elif field == "year_put_in_use":
                value = _coerce_date(value)
            record[field] = value

        extra = {}
        for col_idx, label in extra_col_labels.items():
            value = row[col_idx] if col_idx < len(row) else None
            if isinstance(value, str):
                value = value.strip() or None
            if value not in (None, ""):
                extra[label] = value
        record["extra_fields"] = extra or None

        rows.append(record)
    return rows
